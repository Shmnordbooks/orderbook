#!/usr/bin/env python3
"""
===============================================================
  SHIMANO HARDGOODS CATALOG — Otomatik Güncelleyici
  
  Kullanım:
    python update_catalog.py
  
  Gereksinimler:
    pip install openpyxl
  
  Klasör yapısı:
    update_catalog.py          <- bu script
    index.html                 <- mevcut HTML katalog
    SHIMANO HARDGOODS*.xlsm    <- güncel Excel dosyası
===============================================================
"""

import openpyxl
import json
import re
import os
import sys
import glob
from datetime import datetime

# ─── Ayarlar ────────────────────────────────────────────────
HTML_GLOB    = "index.html"  # Hedef HTML dosya adı
EXCEL_GLOB   = "SHIMANO*.xlsm"  # Excel dosya adı paterni (Boşluklu isimleri kapsar)

# Atlanacak sayfalar — bunlar katalog ürün sayfası değil
SKIP_SHEETS  = {
    'Main Page', 'Stock Availability', 'ConsolidatedOrder',
    'S-Tech Account Request',
    'Groupset', 'Groupset box',
    'Duraace Group', 'Ultegra', '105DI2',
    'New XTRdi2', 'The New XT Di2', 'The New GRX DI2',
    'Dura-Ace R9270', 'Ultegra R8170', '105 DI2 R7100',
    'XTR Di2 M9200', 'Deore XT Di2 M8200', 'GRX Di2 RX820',
}

# Sütun indexleri (0-tabanlı)
COL_GROUP  = 0   # Product Group
COL_CODE   = 1   # Item Number
COL_DESC1  = 2   # Description1
COL_DESC2  = 3   # Description2
COL_QTY    = 4   # QTY
COL_STATUS = 5   # ETD WEEK / Available / N/A
# ────────────────────────────────────────────────────────────

def find_excel():
    """Klasördeki Excel dosyasını bul ve en güncel olanı seç"""
    files = glob.glob(EXCEL_GLOB)
    if not files:
        files = glob.glob("*.xlsm")
    
    # Excel açıkken oluşan gizli kilit dosyalarını (~$) yoksay
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    
    if not files:
        print("❌  Excel dosyası bulunamadı!")
        print(f"   Aranan: {EXCEL_GLOB}")
        sys.exit(1)
        
    # En son değiştirilen dosyayı bul (Böylece aylar değişse de en yeniyi alır)
    files.sort(key=os.path.getmtime, reverse=True)
    
    if len(files) > 1:
        print(f"⚠️   Birden fazla Excel bulundu, en günceli seçildi: {files[0]}")
    else:
        print(f"✅  Excel dosyası bulundu: {files[0]}")
        
    return files[0]

def find_html():
    """Klasördeki HTML katalog dosyasını bul"""
    files = glob.glob(HTML_GLOB)
    if not files:
        files = glob.glob("*.html")
    # Backup dosyalarını hariç tut
    files = [f for f in files if '_backup_' not in f]
    
    if not files:
        print("❌  HTML dosyası bulunamadı!")
        sys.exit(1)
    return files[0]

def format_etd(val):
    """ETD değerini YYYY-MM-DD formatına çevir"""
    if isinstance(val, int):
        s = str(val)
        if len(s) == 8:
            return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()

def parse_status(val):
    """Durum metnini koda çevir"""
    if val is None:
        return "unavailable"
    v = str(val).strip().lower()
    if v == "available":
        return "available"
    elif v in ("n/a", "not available", "unavailable"):
        return "unavailable"
    else:
        return "etd"

def read_excel(path):
    """Excel'den CATALOG dict oluştur"""
    print(f"\n📂  Excel okunuyor: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    
    catalog   = {}
    total     = 0
    skipped   = 0
    
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            skipped += 1
            continue
        
        ws = wb[sheet_name]
        items = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            code  = row[COL_CODE]  if len(row) > COL_CODE  else None
            desc1 = row[COL_DESC1] if len(row) > COL_DESC1 else None
            desc2 = row[COL_DESC2] if len(row) > COL_DESC2 else None
            stat  = row[COL_STATUS] if len(row) > COL_STATUS else None
            
            if not code or not desc1:
                continue
            
            code_str = str(code).strip()
            if code_str.lower().startswith('column'):
                continue
            
            status = parse_status(stat)
            
            item = {
                "code":   code_str,
                "desc1":  str(desc1).strip(),
                "desc2":  str(desc2).strip() if desc2 else "",
                "status": status,
                "etd":    format_etd(stat) if status == "etd" else "",
                "group":  sheet_name
            }
            items.append(item)
        
        if items:
            catalog[sheet_name] = items
            total += len(items)
            print(f"   ✓  {sheet_name:<25} {len(items):>4} ürün")
    
    wb.close()
    print(f"\n   📊  Toplam: {total} ürün, {len(catalog)} kategori")
    return catalog

def get_stats(catalog):
    """İstatistikleri hesapla"""
    avail = etd = unavail = 0
    for items in catalog.values():
        for item in items:
            s = item['status']
            if s == 'available':   avail    += 1
            elif s == 'etd':       etd      += 1
            else:                  unavail += 1
    return avail, etd, unavail

def update_html(catalog, html_file):
    """HTML dosyasını güncelle"""
    print(f"\n🔄  HTML güncelleniyor: {html_file}")
    
    with open(html_file, 'r', encoding='utf-8') as f:
        html = f.read()
    
    pattern = r'^const CATALOG = \{.*\};$'
    new_catalog_js = "const CATALOG = " + json.dumps(catalog, ensure_ascii=False) + ";"
    new_html, count = re.subn(pattern, new_catalog_js, html, count=1, flags=re.MULTILINE)
    
    if count == 0:
        print("❌  HTML içinde 'const CATALOG = {...};' satırı bulunamadı!")
        sys.exit(1)
    
    avail, etd, unavail = get_stats(catalog)
    new_html = re.sub(r'id="headerAvail">[^<]*<', f'id="headerAvail">{avail:,}<', new_html)
    new_html = re.sub(r'id="headerEtd">[^<]*<',   f'id="headerEtd">{etd:,}<',   new_html)
    new_html = re.sub(r'id="headerUnavail">[^<]*<',f'id="headerUnavail">{unavail:,}<', new_html)
    
    total = avail + etd + unavail
    new_html = re.sub(r'(cover-stat-num"[^>]*>)[\d,]+', f'\\g<1>{total:,}', new_html, count=1)
    new_html = re.sub(r'(color:#00c896"[^>]*>)[\d,]+',  f'\\g<1>{avail:,}',  new_html, count=1)
    
    # Yedekleme
    backup = html_file.replace('.html', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.html')
    with open(backup, 'w', encoding='utf-8') as f:
        f.write(html)
    
    # Yeni HTML'i kaydet
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(new_html)
    
    size_kb = os.path.getsize(html_file) // 1024
    print(f"   ✅  Güncellendi! ({size_kb} KB)")
    return avail, etd, unavail, total

def main():
    print("=" * 55)
    print("  SHIMANO HARDGOODS CATALOG — Güncelleyici")
    print("=" * 55)
    
    excel_path = find_excel()
    catalog    = read_excel(excel_path)
    html_file  = find_html()
    avail, etd, unavail, total = update_html(catalog, html_file)
    
    print("\n" + "=" * 55)
    print("  ✅  TAMAMLANDI!")
    print(f"  📦  {total:,} ürün güncellendi")
    print(f"  🟢  Available : {avail:,}")
    print(f"  🟡  ETD       : {etd:,}")
    print(f"  🔴  N/A       : {unavail:,}")
    print("=" * 55)
    print()

if __name__ == "__main__":
    main()